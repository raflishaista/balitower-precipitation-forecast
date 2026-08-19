"""Alpha Vantage stock data collector.

Fetches daily OHLCV for a stock symbol and exports a styled Excel workbook.
"""
import os
import argparse
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def parse_stock_data(api_response):
    if "Time Series (Daily)" not in api_response:
        print("Unexpected API response structure:", list(api_response.keys()))
        return None
    time_series = api_response["Time Series (Daily)"]
    rows = []
    for date_str, values in time_series.items():
        rows.append({
            "Date": date_str,
            "Open": float(values["1. open"]),
            "High": float(values["2. high"]),
            "Low": float(values["3. low"]),
            "Close": float(values["4. close"]),
            "Volume": int(values["5. volume"]),
        })
    df = pd.DataFrame(rows)
    df["Date"] = pd.to_datetime(df["Date"])
    df = df.sort_values("Date").reset_index(drop=True)
    return df


def export_to_excel(df, filename):
    os.makedirs(os.path.dirname(filename) or ".", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Data"

    thin = Side(style="thin", color="AAAAAA")
    thick = Side(style="medium", color="4472C4")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    border = Border(left=thin, right=thin, top=thick, bottom=thick)

    headers = ["Date", "Open", "High", "Low", "Close", "Volume"]
    for col, name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if c_idx in (2, 3, 4, 5):
                cell.number_format = "#,##0.00"
            elif c_idx == 6:
                cell.number_format = "#,##0"

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.freeze_panes = "A2"
    wb.save(filename)
    print(f"Saved to {filename}")


def main():
    parser = argparse.ArgumentParser(description="Alpha Vantage stock data collector")
    parser.add_argument("--symbol", default="IBM", help="Stock symbol (default: IBM)")
    parser.add_argument("--apikey", default="U0PE62MA2UJ0W7LM", help="Alpha Vantage API key")
    args = parser.parse_args()

    url = (f"https://www.alphavantage.co/query?function=TIME_SERIES_DAILY"
           f"&symbol={args.symbol}&apikey={args.apikey}")
    r = requests.get(url)
    data = r.json()

    df = parse_stock_data(data)
    if df is not None:
        print(df.head())
        os.makedirs("data/raw", exist_ok=True)
        export_to_excel(df, f"data/raw/stock_{args.symbol}.xlsx")


if __name__ == "__main__":
    main()
