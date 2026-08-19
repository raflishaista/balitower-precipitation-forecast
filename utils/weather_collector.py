import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta
import time

# NASA POWER Daily Point API - Full Weather Data Collector
# Parameters: Temperature, Precipitation, Humidity, Pressure, Wind Speed/Direction
# Date range: up to ~10+ years (API supports 1981 to near-real-time)
# NOTE: NASA POWER daily point does NOT provide a cloud cover parameter.
#       For cloud cover, use the Monthly or Annual API instead.

PARAMETER_COLS = [
    "Precipitation",    # PRECTOTCORR (mm)
    "Temperature",      # T2M (°C)
    "Temperature Max",  # T2M_MAX (°C)
    "Temperature Min",  # T2M_MIN (°C)
    "Humidity",         # RH2M (%)
    "Pressure",         # PS (kPa)
    "Wind Speed",       # WS10M (m/s)
    "Wind Direction",   # WD10M (degrees)
    "Wind Speed Max",   # WS10M_MAX (m/s)
    "Wind Speed Min",   # WS10M_MIN (m/s)
    "Specific Humidity",# QV2M (g/kg)
    "Surface Temp",     # TS (°C)
    "All-Sky SW Rad",   # ALLSKY_SFC_SW_DWN (W/m²)
    "Clear-Sky SW Rad", # CLRSKY_SFC_SW_DWN (W/m²)
]

SHORT_CODES = ",".join([
    "PRECTOTCORR", "T2M", "T2M_MAX", "T2M_MIN",
    "RH2M", "PS", "WS10M", "WD10M",
    "WS10M_MAX", "WS10M_MIN", "QV2M", "TS",
    "ALLSKY_SFC_SW_DWN", "CLRSKY_SFC_SW_DWN",
])


class NASAPowerWeather:
    def __init__(self):
        self.base_url = "https://power.larc.nasa.gov/api/temporal/daily/point"
        self.session = requests.Session()
        self.session.headers.update({"Accept": "application/json"})

    def fetch(self, latitude, longitude, start_date, end_date):
        params = {
            "parameters": SHORT_CODES,
            "start": start_date,
            "end": end_date,
            "latitude": latitude,
            "longitude": longitude,
            "community": "RE",
            "format": "JSON",
        }
        resp = self.session.get(self.base_url, params=params, timeout=120)
        resp.raise_for_status()
        return self.parse(resp.json())

    def parse(self, api_response):
        if not api_response or "properties" not in api_response:
            return None
        raw = api_response["properties"]["parameter"]
        rows = {}
        for code, values in raw.items():
            if not isinstance(values, dict):
                continue
            for date_str, value in values.items():
                if value is None:
                    continue
                if date_str not in rows:
                    rows[date_str] = {}
                rows[date_str][code] = float(value)
        if not rows:
            return None
        df = pd.DataFrame.from_dict(rows, orient="index")
        df.index.name = "Date"
        df = df.reset_index()
        df["Date"] = pd.to_datetime(df["Date"])
        df = df.sort_values("Date").reset_index(drop=True)
        return df


def export_to_excel(df, filename, location_name, years):
    wb = Workbook()
    ws = wb.active
    ws.title = location_name.replace(" ", "_")[:31]

    thin = Side(style="thin", color="AAAAAA")
    thick = Side(style="medium", color="2E75B6")
    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    border = Border(left=thin, right=thin, top=thick, bottom=thick)

    headers = ["Date"] + PARAMETER_COLS
    for col, name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if c_idx > 1 and isinstance(value, float):
                cell.number_format = "0.00"

    ws.column_dimensions["A"].width = 14
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 14
    ws.freeze_panes = "B2"
    wb.save(filename)
    print(f"  [Saved] {filename}  ({len(df)} days)")


def compute_summary(df, location_name):
    if df is None or df.empty:
        return None
    summary = {"Location": location_name, "Days": len(df),
               "Start": str(df["Date"].min().date()),
               "End": str(df["Date"].max().date())}
    for col in PARAMETER_COLS:
        if col in df.columns:
            vals = df[col].dropna()
            if len(vals) == 0:
                continue
            summary[f"{col}_Sum"] = round(vals.sum(), 2)
            summary[f"{col}_Mean"] = round(vals.mean(), 2)
            summary[f"{col}_Max"] = round(vals.max(), 2)
            summary[f"{col}_Min"] = round(vals.min(), 2)
            summary[f"{col}_Std"] = round(vals.std(), 2)
    return summary


def geocode_location(name):
    """Look up lat/lon for a city name using free Nominatim (OpenStreetMap) API."""
    try:
        resp = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": name, "format": "json", "limit": 1, "accept-language": "en"},
            headers={"User-Agent": "THETA-WeatherCollector/1.0"},
            timeout=15,
        )
        results = resp.json()
        if not results:
            return None
        r = results[0]
        return float(r["lat"]), float(r["lon"]), r.get("display_name", name)
    except Exception:
        return None


def format_display_name(raw_name, query):
    """Build a clean '<City>, <Country>' label from a Nominatim display_name string."""
    parts = [p.strip() for p in raw_name.split(",")]
    if len(parts) >= 2:
        return f"{parts[0]}, {parts[-1]}"
    return query.strip().title()


def lookup_locations(locations_input):
    """Return (name, lat, lon) tuples for each requested location name.
    Uses Nominatim geocoding; falls back to known defaults if lookup fails."""
    known = {
        "jakarta": ("Jakarta, Indonesia", -6.2088, 106.8456),
        "beijing": ("Beijing, China", 39.9042, 116.4074),
        "shanghai": ("Shanghai, China", 31.2304, 121.4737),
        "new york": ("New York, USA", 40.7128, -74.0060),
        "london": ("London, UK", 51.5074, -0.1278),
        "sydney": ("Sydney, Australia", -33.8688, 151.2093),
        "cairo": ("Cairo, Egypt", 30.0444, 31.2357),
        "mumbai": ("Mumbai, India", 19.0760, 72.8777),
    }
    locations = []
    for raw in locations_input:
        key = raw.strip().lower()
        if key in known:
            locations.append(known[key])
            continue
        result = geocode_location(raw.strip())
        if result:
            lat, lon, display = result
            display_name = format_display_name(display, raw)
            locations.append((display_name, lat, lon))
        else:
            print(f"  [WARN] Could not locate '{raw.strip()}'; skipping.")
    return locations


def main():
    default_years = 10
    years_input = input(f"Fetch how many years of data? (default {default_years}): ").strip()
    years = int(years_input) if years_input.isdigit() else default_years

    default_locations = ["jakarta"]
    loc_input = input(f"Which location(s)? (comma-separated, default: {', '.join(default_locations)}): ").strip()
    if loc_input:
        locations_input = [loc.strip() for loc in loc_input.split(",")]
    else:
        locations_input = default_locations

    locations = lookup_locations(locations_input)
    if not locations:
        print("No valid locations provided. Exiting.")
        return

    print(f"\nFetching {years} years of NASA POWER weather data...")
    print(f"Date range: {(datetime.now() - timedelta(days=years*365)).strftime('%Y-%m-%d')} to {datetime.now().strftime('%Y-%m-%d')}")
    print(f"Parameters: {', '.join(PARAMETER_COLS)}")
    print("(Note: cloud cover is not available in the NASA POWER daily point API)")
    print()

    api = NASAPowerWeather()
    all_dfs = {}
    all_summaries = []

    for name, lat, lon in locations:
        print(f"  [{name}] fetching...")
        df = api.fetch(lat, lon,
                       (datetime.now() - timedelta(days=years*365)).strftime("%Y%m%d"),
                       datetime.now().strftime("%Y%m%d"))
        if df is not None:
            all_dfs[name] = df
            all_summaries.append(compute_summary(df, name))
            print(f"    OK: {len(df)} days")
        else:
            print(f"    FAILED")
        time.sleep(0.3)

    for name, df in all_dfs.items():
        safe_name = name.replace(" ", "_")
        export_to_excel(df, f"data/raw/{safe_name}_weather_{years}y.xlsx", name, years)

    if all_summaries:
        sum_df = pd.DataFrame(all_summaries)
        sum_wb = Workbook()
        sum_ws = sum_wb.active
        sum_ws.title = "Summary"

        thin = Side(style="thin", color="AAAAAA")
        thick = Side(style="medium", color="2E75B6")
        header_fill = PatternFill("solid", fgColor="2E75B6")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        border = Border(left=thin, right=thin, top=thick, bottom=thick)

        headers = list(sum_df.columns)
        for col, name in enumerate(headers, 1):
            cell = sum_ws.cell(row=1, column=col, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for r_idx, row in enumerate(dataframe_to_rows(sum_df, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = sum_ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(value, float):
                    cell.number_format = "0.00"

        sum_ws.column_dimensions["A"].width = 20
        for col in range(2, len(headers) + 1):
            sum_ws.column_dimensions[sum_ws.cell(row=1, column=col).column_letter].width = 12
        sum_ws.freeze_panes = "B2"
        os.makedirs("outputs/metrics", exist_ok=True)
        sum_wb.save("outputs/metrics/weather_summary.xlsx")
        print("  [Saved] outputs/metrics/weather_summary.xlsx")

    print("\nDone.")
    sample = next(iter(all_dfs.values()))
    print(f"\nSample (first 5 rows):")
    print(sample.head().to_string())


if __name__ == "__main__":
    main()
