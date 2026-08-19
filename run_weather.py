"""NASA POWER daily weather data collector.

Fetches 14 meteorological parameters (temperature, precipitation, humidity,
pressure, wind, radiation) for one or more cities from the NASA POWER API.
Results are saved as styled Excel workbooks.

Usage:
    python run_weather.py                        # interactive prompts
    python run_weather.py --cities jakarta,cairo --years 5
"""
import argparse
import os
import sys
import time
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import requests


PARAMETER_COLS = [
    "Precipitation", "Temperature", "Temperature Max", "Temperature Min",
    "Humidity", "Pressure", "Wind Speed", "Wind Direction",
    "Wind Speed Max", "Wind Speed Min", "Specific Humidity",
    "Surface Temp", "All-Sky SW Rad", "Clear-Sky SW Rad",
]

SHORT_CODES = ",".join([
    "PRECTOTCORR", "T2M", "T2M_MAX", "T2M_MIN",
    "RH2M", "PS", "WS10M", "WD10M",
    "WS10M_MAX", "WS10M_MIN", "QV2M", "TS",
    "ALLSKY_SFC_SW_DWN", "CLRSKY_SFC_SW_DWN",
])


class NASAPowerWeather:
    def __init__(self):
        self.base_url = "https://power.larc.nasa.gov/api/temporal/daily/point"
        self.session = requests.Session()
        self.session.headers.update({"Accept": "application/json"})

    def fetch(self, latitude, longitude, start_date, end_date):
        params = {
            "parameters": SHORT_CODES,
            "start": start_date,
            "end": end_date,
            "latitude": latitude,
            "longitude": longitude,
            "community": "RE",
            "format": "JSON",
        }
        resp = self.session.get(self.base_url, params=params, timeout=120)
        resp.raise_for_status()
        return self.parse(resp.json())

    def parse(self, api_response):
        if not api_response or "properties" not in api_response:
            return None
        raw = api_response["properties"]["parameter"]
        rows = {}
        for code, values in raw.items():
            if not isinstance(values, dict):
                continue
            for date_str, value in values.items():
                if value is None:
                    continue
                if date_str not in rows:
                    rows[date_str] = {}
                rows[date_str][code] = float(value)
        if not rows:
            return None
        df = pd.DataFrame.from_dict(rows, orient="index")
        df.index.name = "Date"
        df = df.reset_index()
        df["Date"] = pd.to_datetime(df["Date"])
        df = df.sort_values("Date").reset_index(drop=True)
        return df


def geocode_location(name):
    try:
        resp = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": name, "format": "json", "limit": 1, "accept-language": "en"},
            headers={"User-Agent": "THETA-WeatherCollector/1.0"},
            timeout=15,
        )
        results = resp.json()
        if not results:
            return None
        r = results[0]
        return float(r["lat"]), float(r["lon"]), r.get("display_name", name)
    except Exception:
        return None


def format_display_name(raw_name, query):
    parts = [p.strip() for p in raw_name.split(",")]
    if len(parts) >= 2:
        return f"{parts[0]}, {parts[-1]}"
    return query.strip().title()


KNOWN_LOCATIONS = {
    "jakarta": ("Jakarta, Indonesia", -6.2088, 106.8456),
    "beijing": ("Beijing, China", 39.9042, 116.4074),
    "shanghai": ("Shanghai, China", 31.2304, 121.4737),
    "new york": ("New York, USA", 40.7128, -74.0060),
    "london": ("London, UK", 51.5074, -0.1278),
    "sydney": ("Sydney, Australia", -33.8688, 151.2093),
    "cairo": ("Cairo, Egypt", 30.0444, 31.2357),
    "mumbai": ("Mumbai, India", 19.0760, 72.8777),
}


def lookup_locations(locations_input):
    locations = []
    for raw in locations_input:
        key = raw.strip().lower()
        if key in KNOWN_LOCATIONS:
            locations.append(KNOWN_LOCATIONS[key])
            continue
        result = geocode_location(raw.strip())
        if result:
            lat, lon, display = result
            display_name = format_display_name(display, raw)
            locations.append((display_name, lat, lon))
        else:
            print(f"  [WARN] Could not locate '{raw.strip()}'; skipping.")
    return locations


def export_to_excel(df, filename, location_name):
    os.makedirs(os.path.dirname(filename) or ".", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = location_name.replace(" ", "_")[:31]

    thin = Side(style="thin", color="AAAAAA")
    thick = Side(style="medium", color="2E75B6")
    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    border = Border(left=thin, right=thin, top=thick, bottom=thick)

    headers = ["Date"] + PARAMETER_COLS
    for col, name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if c_idx > 1 and isinstance(value, float):
                cell.number_format = "0.00"

    ws.column_dimensions["A"].width = 14
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 14
    ws.freeze_panes = "B2"
    wb.save(filename)
    print(f"  [Saved] {filename}  ({len(df)} days)")


def compute_summary(df, location_name):
    if df is None or df.empty:
        return None
    summary = {"Location": location_name, "Days": len(df),
               "Start": str(df["Date"].min().date()),
               "End": str(df["Date"].max().date())}
    for col in PARAMETER_COLS:
        if col in df.columns:
            vals = df[col].dropna()
            if len(vals) == 0:
                continue
            summary[f"{col}_Sum"] = round(vals.sum(), 2)
            summary[f"{col}_Mean"] = round(vals.mean(), 2)
            summary[f"{col}_Max"] = round(vals.max(), 2)
            summary[f"{col}_Min"] = round(vals.min(), 2)
            summary[f"{col}_Std"] = round(vals.std(), 2)
    return summary


def main():
    parser = argparse.ArgumentParser(description="NASA POWER weather data collector")
    parser.add_argument("--cities", default="jakarta",
                        help="Comma-separated city names (default: jakarta)")
    parser.add_argument("--years", type=int, default=10,
                        help="Years of historical data (default: 10)")
    args = parser.parse_args()

    locations_input = [loc.strip() for loc in args.cities.split(",")]
    locations = lookup_locations(locations_input)
    if not locations:
        print("No valid locations provided. Exiting.")
        return

    years = args.years
    start_date = (datetime.now() - timedelta(days=years * 365)).strftime("%Y%m%d")
    end_date = datetime.now().strftime("%Y%m%d")

    print(f"\nFetching {years} years of NASA POWER weather data...")
    print(f"Date range: {start_date} to {end_date}")
    print(f"Parameters: {', '.join(PARAMETER_COLS)}")
    print()

    api = NASAPowerWeather()
    all_dfs = {}
    all_summaries = []

    for name, lat, lon in locations:
        print(f"  [{name}] fetching...")
        df = api.fetch(lat, lon, start_date, end_date)
        if df is not None:
            all_dfs[name] = df
            all_summaries.append(compute_summary(df, name))
            print(f"    OK: {len(df)} days")
        else:
            print(f"    FAILED")
        time.sleep(0.3)

    for name, df in all_dfs.items():
        safe_name = name.replace(" ", "_")
        export_to_excel(df, f"data/raw/{safe_name}_weather_{years}y.xlsx", name)

    if all_summaries:
        sum_df = pd.DataFrame(all_summaries)
        sum_wb = Workbook()
        sum_ws = sum_wb.active
        sum_ws.title = "Summary"

        thin = Side(style="thin", color="AAAAAA")
        thick = Side(style="medium", color="2E75B6")
        header_fill = PatternFill("solid", fgColor="2E75B6")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        border = Border(left=thin, right=thin, top=thick, bottom=thick)

        headers = list(sum_df.columns)
        for col, name in enumerate(headers, 1):
            cell = sum_ws.cell(row=1, column=col, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for r_idx, row in enumerate(dataframe_to_rows(sum_df, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = sum_ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(value, float):
                    cell.number_format = "0.00"

        sum_ws.column_dimensions["A"].width = 20
        for col in range(2, len(headers) + 1):
            sum_ws.column_dimensions[sum_ws.cell(row=1, column=col).column_letter].width = 12
        sum_ws.freeze_panes = "B2"
        os.makedirs("outputs/metrics", exist_ok=True)
        sum_wb.save("outputs/metrics/weather_summary.xlsx")
        print("  [Saved] outputs/metrics/weather_summary.xlsx")

    # Export each city as JSON for use by precipitation_forecast.py
    import json
    os.makedirs("inputs_json", exist_ok=True)
    # NASA POWER code -> friendly name mapping
    NAASA_POWER_TO_JSON = {
        "PRECTOTCORR": "precip_mm",
        "RH2M": "humidity",
        "PS": "pressure",
        "WS10M": "windspeed",
        "T2M": "temperature",
        "WD10M": "wind_direction",
    }
    for name, df in all_dfs.items():
        safe = name.replace(" ", "_")
        out = {"city": name, "lat": None, "lon": None, "years": years, "dates": []}
        for _, row in df.iterrows():
            date_str = row["Date"].strftime("%Y-%m-%d")
            out["dates"].append(date_str)
            for code, key in NAASA_POWER_TO_JSON.items():
                if code in row and row[code] is not None:
                    if key not in out:
                        out[key] = []
                    out[key].append(round(float(row[code]), 4))
        with open(f"inputs_json/{safe}_precip.json", "w", encoding="utf-8") as f:
            json.dump(out, f, indent=2)
        print(f"  [Saved] inputs_json/{safe}_precip.json")

    print("\nDone.")
    sample = next(iter(all_dfs.values()))
    print(f"\nSample (first 5 rows):")
    print(sample.head().to_string())


if __name__ == "__main__":
    main()
